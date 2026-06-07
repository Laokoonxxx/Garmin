"""Same as heatmap2_2026 but for year 2025."""
import json
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

RUN_TAGS = {"running", "run", "9", "trail_running", "trail running", "treadmill_running"}
IN = r"C:\Users\mares\Desktop\garmin\activities.json"
OUT_XLSX = r"C:\Users\mares\Desktop\garmin\plan_2025_behy.xlsx"
YEAR = 2025
GOAL_KM = 1000
COLS = 50; ROWS = 20

PACE_TIERS = [
    ("světle zelená", "> 7:00",       420,   100000, "B5E61D"),
    ("zelená",        "6:30–7:00",    390,      420, "2E8B2E"),
    ("žlutá",         "6:00–6:30",    360,      390, "FFD700"),
    ("oranžová",      "5:30–6:00",    330,      360, "FF8C00"),
    ("červená",       "5:00–5:30",    300,      330, "E81416"),
    ("modrá",         "4:30–5:00",    270,      300, "2E70F0"),
    ("fialová",       "4:15–4:30",    255,      270, "8A2BE2"),
    ("růžová",        "< 4:15",         0,      255, "FF69B4"),
]
EMPTY_FILL = "FFFFFF"


def tier_for(pace_s):
    for i, (_, _, lo, hi, _) in enumerate(PACE_TIERS):
        if lo <= pace_s < hi:
            return i
    return 0


def main():
    acts = json.load(open(IN, encoding="utf-8"))
    valid = []
    for a in acts:
        sport = (a.get("sport") or "").strip().lower()
        if sport not in RUN_TAGS:
            continue
        dist_m = a["dist_m"]; moving_s = a["moving_s"] or 0
        if dist_m > 2000 and moving_s > 300:
            pace = (moving_s/60.0) / (dist_m/1000.0)
            if pace < 4.5 or pace > 10.0:
                continue
        dt = datetime.fromisoformat(a["start"])
        if dt.year != YEAR:
            continue
        valid.append((dt, a))
    valid.sort(key=lambda x: x[0])

    runs = []
    cum_km = 0.0
    cum_cells = 0
    tier_cells = [0] * len(PACE_TIERS)
    for dt, a in valid:
        d_km = a["dist_m"] / 1000.0
        pace_s = (a["moving_s"] / d_km) if a["moving_s"] and d_km > 0 else 0
        cum_km += d_km
        new_cum_cells = int(round(cum_km))
        n_cells = new_cum_cells - cum_cells
        cum_cells = new_cum_cells
        if n_cells <= 0:
            continue
        ti = tier_for(pace_s)
        tier_cells[ti] += n_cells
        runs.append((n_cells, pace_s, dt, d_km, ti))

    colored_n = sum(r[0] for r in runs)
    print(f"{YEAR} runs: {len(valid)}; total km {cum_km:.2f} -> {colored_n} cells")

    cell_paint = []
    for n, pace_s, dt, d_km, ti in runs:
        color = PACE_TIERS[ti][4]
        m = int(pace_s // 60); s = int(round(pace_s - m*60))
        tip = f"{dt.strftime('%Y-%m-%d')}  {d_km:.2f} km  {m}:{s:02d}/km"
        for k in range(n):
            cell_paint.append((color, tip, k == n - 1))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Plán {YEAR} — běhy"
    thin = Side(border_style="thin", color="DDDDDD")
    thick_black = Side(border_style="thick", color="000000")
    center = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1,
            value=f"Rok {YEAR} — 1000 km (běh = blok, tlustá čára = konec běhu)")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COLS)
    ws.cell(row=2, column=1,
            value=f"Uběhnuto: {colored_n} km z {GOAL_KM} km "
                  f"({colored_n/GOAL_KM*100:.1f}%), {len(runs)} běhů")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=COLS)

    GRID_TOP = 4

    for i in range(GOAL_KM):
        r = GRID_TOP + (i // COLS)
        c = 1 + (i % COLS)
        cell = ws.cell(row=r, column=c, value=i + 1)
        cell.alignment = center
        cell.font = Font(size=8, color="666666")

        left = thin; right = thin; top = thin; bottom = thin
        if i < len(cell_paint):
            color, tip, is_last = cell_paint[i]
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(size=8, color="FFFFFF", bold=True)
            cell.comment = Comment(tip, "stats")
            if is_last and i != len(cell_paint) - 1:
                if c < COLS:
                    right = thick_black
                else:
                    bottom = thick_black
        else:
            cell.fill = PatternFill("solid", fgColor=EMPTY_FILL)
        cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    for i in range(1, GOAL_KM):
        if i - 1 < len(cell_paint) and cell_paint[i - 1][2] and i - 1 != len(cell_paint) - 1:
            c_prev = 1 + ((i - 1) % COLS)
            r_cur = GRID_TOP + (i // COLS)
            c_cur = 1 + (i % COLS)
            cell_cur = ws.cell(row=r_cur, column=c_cur)
            b = cell_cur.border
            if c_prev == COLS:
                cell_cur.border = Border(left=b.left, right=b.right,
                                          top=thick_black, bottom=b.bottom)
            else:
                cell_cur.border = Border(left=thick_black, right=b.right,
                                          top=b.top, bottom=b.bottom)

    for c in range(1, COLS + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 4
    for r in range(GRID_TOP, GRID_TOP + ROWS):
        ws.row_dimensions[r].height = 20

    legend_row = GRID_TOP + ROWS + 2
    ws.cell(row=legend_row, column=1,
            value="Legenda (průměrné tempo běhu, pomalé → rychlé):").font = Font(bold=True)
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, (name, label, lo, hi, hex_color) in enumerate(PACE_TIERS):
        rr = legend_row + 1 + i
        sw = ws.cell(row=rr, column=1, value="")
        sw.fill = PatternFill("solid", fgColor=hex_color); sw.border = border_thin
        ws.cell(row=rr, column=2, value=name)
        ws.cell(row=rr, column=3, value=label)
        ws.cell(row=rr, column=4, value=f"{tier_cells[i]} km")
    rr = legend_row + 1 + len(PACE_TIERS)
    sw = ws.cell(row=rr, column=1, value="")
    sw.fill = PatternFill("solid", fgColor=EMPTY_FILL); sw.border = border_thin
    ws.cell(row=rr, column=2, value="zbývá")
    ws.cell(row=rr, column=4, value=f"{GOAL_KM - colored_n} km")

    list_row = legend_row + len(PACE_TIERS) + 4
    ws.cell(row=list_row, column=1, value="Seznam běhů:").font = Font(bold=True)
    hdr = ["datum", "km", "tempo", "buněk", "barva"]
    for j, h in enumerate(hdr):
        ws.cell(row=list_row + 1, column=1 + j, value=h).font = Font(bold=True)
    for idx, (n, pace_s, dt, d_km, ti) in enumerate(runs):
        rr = list_row + 2 + idx
        m = int(pace_s // 60); s = int(round(pace_s - m*60))
        ws.cell(row=rr, column=1, value=dt.strftime("%Y-%m-%d"))
        ws.cell(row=rr, column=2, value=round(d_km, 2))
        ws.cell(row=rr, column=3, value=f"{m}:{s:02d}/km")
        ws.cell(row=rr, column=4, value=n)
        sw = ws.cell(row=rr, column=5, value=PACE_TIERS[ti][0])
        sw.fill = PatternFill("solid", fgColor=PACE_TIERS[ti][4])

    wb.save(OUT_XLSX)
    print(f"Saved: {OUT_XLSX}")
    for (name, label, _, _, _), cnt in zip(PACE_TIERS, tier_cells):
        print(f"  {name:14s} {label:11s} {cnt} km")


if __name__ == "__main__":
    main()
