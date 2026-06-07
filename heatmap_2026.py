"""Generate 1000-cell Excel heatmap for 2026 goal.
- 1000 cells (50 cols x 20 rows) = 1000 km goal.
- First N cells (N = total km run in 2026 so far) are colored by pace of that km.
- 5 pace tiers, green (fast) -> red (slow).
"""
import json
import importlib.util
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

spec = importlib.util.spec_from_file_location("analyze", r"C:\Users\mares\Desktop\garmin\analyze.py")
mod = importlib.util.module_from_spec(spec); spec.loader.exec_module(mod)
clean_series = mod.clean_series

RUN_TAGS = {"running", "run", "9", "trail_running", "trail running", "treadmill_running"}
IN = r"C:\Users\mares\Desktop\garmin\activities.json"
OUT_XLSX = r"C:\Users\mares\Desktop\garmin\plan_2026.xlsx"
YEAR = 2026
GOAL_KM = 1000

# 5 tiers, fastest -> slowest, green -> red
PACE_TIERS = [
    ("≤ 5:00",       0,       300, "1A9850"),   # dark green
    ("5:00–5:30",  300,       330, "66BD63"),   # light green
    ("5:30–6:00",  330,       360, "FEE08B"),   # yellow
    ("6:00–7:00",  360,       420, "FDAE61"),   # orange
    ("> 7:00",     420,    100000, "D73027"),   # red
]
NO_DATA_FILL = "BFBFBF"    # gray for kms without pace info
EMPTY_FILL = "FFFFFF"      # white for remaining goal


def km_splits(series):
    """Yield pace (s/km) for each completed km in cleaned series."""
    if len(series) < 2:
        return
    target = 1000.0
    t_prev_km = series[0][0]
    next_mark = series[0][1] + target
    for i in range(1, len(series)):
        while next_mark <= series[i][1]:
            t_a, d_a = series[i-1]
            t_b, d_b = series[i]
            if d_b <= d_a:
                t_at = t_b
            else:
                frac = (next_mark - d_a) / (d_b - d_a)
                frac = max(0.0, min(1.0, frac))
                t_at = t_a + frac * (t_b - t_a)
            pace_s = t_at - t_prev_km
            if 150 < pace_s < 1200:   # plausible pace 2:30–20:00 min/km
                yield pace_s
            t_prev_km = t_at
            next_mark += target


def main():
    acts = json.load(open(IN, encoding="utf-8"))
    # Filter 2026 running activities (same filters as analyze.py)
    valid = []
    for a in acts:
        sport = (a.get("sport") or "").strip().lower()
        if sport not in RUN_TAGS:
            continue
        dist_m = a["dist_m"]; moving_s = a["moving_s"] or 0
        if dist_m > 2000 and moving_s > 300:
            pace = (moving_s/60.0) / (dist_m/1000.0)
            if pace < 4.5 or pace > 10.0:
                continue
        dt = datetime.fromisoformat(a["start"])
        if dt.year != YEAR:
            continue
        valid.append((dt, a))

    valid.sort(key=lambda x: x[0])
    print(f"2026 running activities: {len(valid)}")

    # Build chronological list of (pace_s, source_act_idx). Also keep "extra" partial km
    # accounted as one final split at average pace.
    splits = []
    total_km = 0.0
    for dt, a in valid:
        raw = [(float(t), float(d)) for t, d in (a.get("series") or [])]
        s = clean_series(raw)
        per_km = list(km_splits(s))
        for p in per_km:
            splits.append(p)
        # account for leftover (< 1km tail) using activity-average pace
        total_km += a["dist_m"] / 1000.0
    # round count of colored cells to match total_km
    colored_n = int(round(total_km))
    # if we have fewer per-km splits than colored_n, pad with average paces
    if splits:
        avg = sum(splits) / len(splits)
    else:
        avg = 360
    while len(splits) < colored_n:
        splits.append(avg)
    splits = splits[:colored_n]
    print(f"Total km: {total_km:.2f} -> colored cells: {colored_n}, pace samples: {len(splits)}")

    # Bucket counts
    bucket_counts = [0]*len(PACE_TIERS)
    for p in splits:
        for idx, (_, lo, hi, _) in enumerate(PACE_TIERS):
            if lo <= p < hi:
                bucket_counts[idx] += 1
                break

    # ---- Build workbook ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plán 2026"

    COLS = 50; ROWS = 20  # 1000 cells
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # Title row
    ws.cell(row=1, column=1, value=f"Plán 2026 — 1000 km")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COLS)
    ws.cell(row=2, column=1, value=f"Uběhnuto: {colored_n} km z {GOAL_KM} km ({colored_n/GOAL_KM*100:.1f}%)")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=COLS)

    GRID_TOP = 4
    for i in range(GOAL_KM):
        r = GRID_TOP + (i // COLS)
        c = 1 + (i % COLS)
        cell = ws.cell(row=r, column=c, value=i + 1)
        cell.alignment = center
        cell.border = border
        cell.font = Font(size=8, color="666666")
        if i < colored_n:
            p = splits[i]
            color = None
            for label, lo, hi, hex_color in PACE_TIERS:
                if lo <= p < hi:
                    color = hex_color
                    break
            if color is None:
                color = NO_DATA_FILL
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(size=8, color="FFFFFF", bold=True)
        else:
            cell.fill = PatternFill("solid", fgColor=EMPTY_FILL)

    # Column / row sizes
    for c in range(1, COLS+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 4
    for r in range(GRID_TOP, GRID_TOP + ROWS):
        ws.row_dimensions[r].height = 20

    # Legend
    legend_row = GRID_TOP + ROWS + 2
    ws.cell(row=legend_row, column=1, value="Legenda (tempo min/km):").font = Font(bold=True)
    for i, (label, lo, hi, hex_color) in enumerate(PACE_TIERS):
        rr = legend_row + 1 + i
        sw = ws.cell(row=rr, column=1, value="")
        sw.fill = PatternFill("solid", fgColor=hex_color)
        sw.border = border
        ws.cell(row=rr, column=2, value=label)
        ws.cell(row=rr, column=3, value=f"{bucket_counts[i]} km")
    # Empty/no data swatches
    rr = legend_row + 1 + len(PACE_TIERS)
    sw = ws.cell(row=rr, column=1, value=""); sw.fill = PatternFill("solid", fgColor=EMPTY_FILL); sw.border = border
    ws.cell(row=rr, column=2, value="zbývá uběhnout")
    ws.cell(row=rr, column=3, value=f"{GOAL_KM - colored_n} km")

    # Summary
    summ_row = legend_row + 1 + len(PACE_TIERS) + 3
    ws.cell(row=summ_row, column=1, value="Souhrn").font = Font(bold=True)
    ws.cell(row=summ_row+1, column=1, value="Aktivit v 2026:")
    ws.cell(row=summ_row+1, column=2, value=len(valid))
    if splits:
        avg_pace = sum(splits) / len(splits)
        m = int(avg_pace // 60); s = int(round(avg_pace - m*60))
        ws.cell(row=summ_row+2, column=1, value="Průměrné tempo:")
        ws.cell(row=summ_row+2, column=2, value=f"{m}:{s:02d} min/km")

    wb.save(OUT_XLSX)
    print(f"Saved: {OUT_XLSX}")
    print("Distribution:")
    for (label, _, _, _), cnt in zip(PACE_TIERS, bucket_counts):
        print(f"  {label}: {cnt} km")


if __name__ == "__main__":
    main()
